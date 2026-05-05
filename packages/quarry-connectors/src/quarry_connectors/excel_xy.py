"""ExcelXYConnector — materializes Excel files (.xlsx, .xls) with coordinate columns.

Lane: connector

This connector auto-detects lat/lon columns in Excel spreadsheets and converts
tabular data to vector artifacts. Supports both eager (GeoPackage output) and
lazy (metadata-only) materialization.

Key features:
- Auto-detect coordinate columns (lat/lon, latitude/longitude, x/y, easting/northing)
- Sheet selection via "path/to/file.xlsx::sheet_name" syntax
- Explicit column specification via "path/to/file.xlsx::sheet_name::lon_col,lat_col"
- CRS inference based on column names
- Skips rows with empty or non-numeric coordinates
- Falls back to TABLE artifact type when no coordinate columns detected

Dependencies:
- openpyxl (optional): for reading .xlsx files
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

import fiona
from fiona.crs import CRS
from quarry_core.artifact import (
    Artifact,
    ArtifactType,
    BackingStore,
    BackingStoreKind,
    Lineage,
    SpatialDescriptor,
    content_hash,
)
from quarry_core.connector import (
    CatalogEntry,
    ConnectorCapability,
    MaterializeError,
    MaterializeResult,
)

if TYPE_CHECKING:
    from quarry_core.source_ref import SourceRef

# Optional dependency: openpyxl
try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

# Column name patterns for coordinate detection (case-insensitive)
LON_PATTERNS = frozenset({"lon", "longitude", "lng", "long", "x", "easting"})
LAT_PATTERNS = frozenset({"lat", "latitude", "y", "northing"})


def _normalize_column(name: str) -> str:
    """Normalize column name for pattern matching."""
    return name.lower().strip().replace("_", "").replace("-", "")


def _detect_coordinate_columns(
    headers: list[str],
) -> tuple[str | None, str | None, str | None]:
    """Detect coordinate columns from headers.

    Returns:
        Tuple of (lon_col, lat_col, crs_hint) where crs_hint is "EPSG:4326" for lat/lon
        or None for projected coordinates.
    """
    lon_col: str | None = None
    lat_col: str | None = None
    crs_hint: str | None = None

    for header in headers:
        normalized = _normalize_column(header)

        if normalized in LON_PATTERNS:
            lon_col = header
            # Check if this is lat/lon (WGS84) or x/y (projected)
            if normalized in ("lon", "longitude", "lng", "long"):
                crs_hint = "EPSG:4326"
        elif normalized in LAT_PATTERNS:
            lat_col = header
            if normalized in ("lat", "latitude"):
                crs_hint = "EPSG:4326"

    return lon_col, lat_col, crs_hint


def _parse_source_ref(
    source_ref: SourceRef | str,
) -> tuple[Path, str | None, str | None, str | None]:
    """Parse source_ref into (file_path, sheet_name, explicit_lon_col, explicit_lat_col).

    Supports:
    - "path/to/file.xlsx" → auto-detect sheet (first) and columns
    - "path/to/file.xlsx::sheet_name" → specific sheet, auto-detect columns
    - "path/to/file.xlsx::sheet_name::lon_col,lat_col" → specific sheet + explicit columns
    """
    raw = str(source_ref).strip()

    # Check for double :: separator (sheet + columns)
    if "::" in raw:
        parts = raw.split("::")
        if len(parts) == 3:
            # path::sheet_name::lon_col,lat_col
            path_part, sheet_name, col_part = parts
            if "," in col_part:
                lon_col, lat_col = col_part.split(",", 1)
                return Path(path_part), sheet_name, lon_col.strip(), lat_col.strip()
            # Invalid format, treat as path::sheet_name
            return Path(path_part), sheet_name, None, None
        elif len(parts) == 2:
            # path::sheet_name
            path_part, sheet_name = parts
            return Path(path_part), sheet_name, None, None

    return Path(raw), None, None, None


def _is_numeric(value: Any) -> bool:
    """Check if a value is numeric (int or float)."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        if not value.strip():
            return False
        try:
            float(value)
            return True
        except ValueError:
            return False
    return False


def _to_float(value: Any) -> float | None:
    """Convert a value to float, returning None if not possible."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return None
    return None


def _open_workbook_and_sheet(
    file_path: Path,
    sheet_name: str | None,
) -> tuple[Any, Any, str, list[str]]:
    """Validate file existence and open Excel workbook, selecting the requested sheet.

    Returns:
        Tuple of (workbook, worksheet, resolved_sheet_name, sheet_names)
    """
    if not HAS_OPENPYXL:
        raise MaterializeError(
            str(file_path),
            "openpyxl is required for Excel support. Install with: pip install openpyxl",
        )

    if not file_path.exists():
        raise MaterializeError(str(file_path), f"File not found: {file_path}")

    if file_path.stat().st_size == 0:
        raise MaterializeError(str(file_path), "File is empty")

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        raise MaterializeError(str(file_path), f"Failed to open Excel file: {e}") from e

    sheet_names = wb.sheetnames

    if sheet_name is not None:
        if sheet_name not in sheet_names:
            wb.close()
            raise MaterializeError(
                str(file_path), f"Sheet not found: '{sheet_name}'. Available: {sheet_names}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    return wb, ws, sheet_name, sheet_names


def _read_header_row(wb: Any, ws: Any, file_path: Path) -> list[str]:
    """Read the header row from the worksheet."""
    try:
        headers = []
        for cell in next(ws.rows):
            headers.append(str(cell.value) if cell.value is not None else "")
    except StopIteration:
        wb.close()
        raise MaterializeError(str(file_path), "Sheet has no header row")
    return headers


def _resolve_coordinate_columns(
    headers: list[str],
    explicit_lon_col: str | None,
    explicit_lat_col: str | None,
    wb: Any,
    file_path: Path,
) -> tuple[str | None, str | None, str | None]:
    """Resolve longitude and latitude columns.

    Returns:
        Tuple of (lon_col, lat_col, crs_hint)
    """
    lon_col: str | None = explicit_lon_col
    lat_col: str | None = explicit_lat_col
    crs_hint: str | None = None

    if lon_col is None or lat_col is None:
        detected_lon, detected_lat, detected_crs = _detect_coordinate_columns(headers)
        if lon_col is None:
            lon_col = detected_lon
        if lat_col is None:
            lat_col = detected_lat
        crs_hint = detected_crs
    else:
        if lon_col not in headers:
            wb.close()
            raise MaterializeError(
                str(file_path), f"Explicit longitude column not found: {lon_col}"
            )
        if lat_col not in headers:
            wb.close()
            raise MaterializeError(str(file_path), f"Explicit latitude column not found: {lat_col}")
        norm_lon = _normalize_column(lon_col)
        norm_lat = _normalize_column(lat_col)
        if norm_lon in ("lon", "longitude", "lng", "long") or norm_lat in (
            "lat",
            "latitude",
        ):
            crs_hint = "EPSG:4326"

    return lon_col, lat_col, crs_hint


def _scan_data_rows(
    ws: Any,
    headers: list[str],
    lon_col: str | None,
    lat_col: str | None,
    wb: Any,
    file_path: Path,
) -> tuple[int, int, tuple[float, float, float, float] | None]:
    """Scan data rows to count rows and calculate extent.

    Returns:
        Tuple of (row_count, valid_feature_count, extent)
    """
    row_count = 0
    valid_feature_count = 0
    extent: tuple[float, float, float, float] | None = None
    x_values: list[float] = []
    y_values: list[float] = []

    has_coordinates = lon_col is not None and lat_col is not None

    if has_coordinates:
        try:
            lon_idx = headers.index(lon_col)
            lat_idx = headers.index(lat_col)
        except ValueError:
            wb.close()
            raise MaterializeError(
                str(file_path), f"Coordinate column not found in headers: {lon_col}, {lat_col}"
            )

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1

            if len(row) <= max(lon_idx, lat_idx):
                continue

            lon_val = row[lon_idx]
            lat_val = row[lat_idx]

            if _is_numeric(lon_val) and _is_numeric(lat_val):
                x = _to_float(lon_val)
                y = _to_float(lat_val)
                if x is not None and y is not None:
                    x_values.append(x)
                    y_values.append(y)
                    valid_feature_count += 1
    else:
        for _ in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1

    if x_values and y_values:
        extent = (min(x_values), min(y_values), max(x_values), max(y_values))

    return row_count, valid_feature_count, extent


def _read_excel_metadata(
    file_path: Path,
    sheet_name: str | None = None,
    explicit_lon_col: str | None = None,
    explicit_lat_col: str | None = None,
) -> dict[str, Any]:
    """Read metadata from an Excel file.

    Returns dict with:
    - columns: list of column names
    - row_count: number of data rows
    - sheet_name: sheet that was read
    - sheet_names: list of all sheet names in the workbook
    - lon_col: detected or explicit longitude column
    - lat_col: detected or explicit latitude column
    - crs_hint: "EPSG:4326" or None
    - has_coordinates: whether coordinate columns were found
    - valid_feature_count: count of rows with valid numeric coordinates
    - extent: (xmin, ymin, xmax, ymax) or None
    """
    wb, ws, resolved_sheet_name, sheet_names = _open_workbook_and_sheet(file_path, sheet_name)
    headers = _read_header_row(wb, ws, file_path)
    lon_col, lat_col, crs_hint = _resolve_coordinate_columns(
        headers, explicit_lon_col, explicit_lat_col, wb, file_path
    )
    has_coordinates = lon_col is not None and lat_col is not None
    row_count, valid_feature_count, extent = _scan_data_rows(
        ws, headers, lon_col, lat_col, wb, file_path
    )

    wb.close()

    return {
        "columns": headers,
        "row_count": row_count,
        "sheet_name": resolved_sheet_name,
        "sheet_names": sheet_names,
        "lon_col": lon_col,
        "lat_col": lat_col,
        "crs_hint": crs_hint,
        "has_coordinates": has_coordinates,
        "valid_feature_count": valid_feature_count,
        "extent": extent,
    }


def _write_geopackage(
    input_path: Path,
    output_path: Path,
    metadata: dict[str, Any],
    default_crs: str = "EPSG:4326",
) -> None:
    """Write Excel data to GeoPackage using fiona.

    Builds shapely Points from coordinate columns and writes via fiona.
    Skips rows with empty or non-numeric coordinates.
    """
    if not HAS_OPENPYXL:
        raise MaterializeError(
            str(input_path),
            "openpyxl is required for Excel support. Install with: pip install openpyxl",
        )

    headers = metadata["columns"]
    sheet_name = metadata["sheet_name"]
    lon_col = metadata["lon_col"]
    lat_col = metadata["lat_col"]
    crs_hint = metadata["crs_hint"]

    lon_idx = headers.index(lon_col)
    lat_idx = headers.index(lat_col)

    # Determine CRS
    crs = crs_hint if crs_hint is not None else default_crs

    # Build property schema (all columns except coordinates)
    prop_schema = {}
    for i, header in enumerate(headers):
        if header == lon_col or header == lat_col:
            continue
        # Default to string type for all properties
        prop_schema[header] = "str"

    schema = {"geometry": "Point", "properties": prop_schema}

    # Parse CRS
    if crs.startswith("EPSG:"):
        crs_obj = CRS.from_epsg(int(crs.split(":")[1]))
    else:
        crs_obj = CRS.from_epsg(4326)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with fiona.open(
        str(output_path),
        "w",
        driver="GPKG",
        schema=schema,
        crs=crs_obj,
    ) as dst:
        wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)
        ws = wb[sheet_name]

        # Skip header row
        first_row = True
        for row in ws.iter_rows(values_only=True):
            if first_row:
                first_row = False
                continue

            if len(row) <= max(lon_idx, lat_idx):
                continue

            lon_val = row[lon_idx]
            lat_val = row[lat_idx]

            if not _is_numeric(lon_val) or not _is_numeric(lat_val):
                continue

            x = _to_float(lon_val)
            y = _to_float(lat_val)
            if x is None or y is None:
                continue

            # Build properties dict
            props = {}
            for i, header in enumerate(headers):
                if header == lon_col or header == lat_col:
                    continue
                if i < len(row):
                    val = row[i]
                    props[header] = str(val) if val is not None else ""
                else:
                    props[header] = ""

            # Create feature with GeoJSON geometry dict
            feature = {
                "geometry": {
                    "type": "Point",
                    "coordinates": [x, y],
                },
                "properties": props,
            }
            dst.write(feature)

        wb.close()


class ExcelXYConnector:
    """Materializes Excel files with coordinate columns into canonical artifacts.

    Auto-detects lat/lon columns and converts tabular data to vector artifacts.
    Supports explicit sheet and column specification via
    "path/to/file.xlsx::sheet_name::lon_col,lat_col" syntax.
    """

    def __init__(self, default_crs: str = "EPSG:4326"):
        """Initialize connector with optional default CRS for projected coordinates.

        Args:
            default_crs: CRS to use for x/y/easting/northing columns (default: EPSG:4326)
        """
        self.default_crs = default_crs

    @property
    def name(self) -> str:
        return "excel_xy"

    @property
    def capabilities(self) -> ConnectorCapability:
        return (
            ConnectorCapability.MATERIALIZE
            | ConnectorCapability.DISCOVER
            | ConnectorCapability.MATERIALIZE_LAZY
            | ConnectorCapability.METADATA_ONLY
        )

    def materialize(
        self,
        source_ref: SourceRef | str,
        workspace: Path,
        *,
        lazy: bool = False,
    ) -> MaterializeResult:
        """Materialize an Excel file into a canonical artifact.

        Args:
            source_ref: Path to Excel file, optionally with "::sheet_name" or
                       "::sheet_name::lon_col,lat_col" suffix
            workspace: Where to write output GeoPackage (eager mode)
            lazy: If True, return metadata-only artifact with LAZY_HANDLE backing

        Returns:
            MaterializeResult with the artifact and provenance.

        Raises:
            MaterializeError: If materialization fails.
        """
        file_path, sheet_name, explicit_lon_col, explicit_lat_col = _parse_source_ref(source_ref)

        # Read metadata
        try:
            meta = _read_excel_metadata(
                file_path,
                sheet_name=sheet_name,
                explicit_lon_col=explicit_lon_col,
                explicit_lat_col=explicit_lat_col,
            )
        except MaterializeError:
            raise
        except Exception as e:
            raise MaterializeError(source_ref, f"Failed to read Excel metadata: {e}") from e

        # Determine artifact type and CRS
        has_coords = meta["has_coordinates"]
        crs = meta["crs_hint"] if meta["crs_hint"] is not None else self.default_crs

        # Build spatial descriptor
        if has_coords:
            spatial = SpatialDescriptor(
                crs=crs,
                extent=meta["extent"],
                feature_count=meta["valid_feature_count"],
            )
            artifact_type = ArtifactType.VECTOR
        else:
            spatial = SpatialDescriptor(
                crs=None,
                extent=None,
                feature_count=meta["row_count"],
            )
            artifact_type = ArtifactType.TABLE

        # Build lineage params
        lineage_params: dict[str, Any] = {
            "source": "excel_xy",
            "path": str(file_path),
            "sheet_name": meta["sheet_name"],
            "sheet_names": meta["sheet_names"],
            "lazy": lazy,
            "columns": meta["columns"],
            "row_count": meta["row_count"],
            "detected_lon_col": meta["lon_col"],
            "detected_lat_col": meta["lat_col"],
            "has_coordinates": has_coords,
        }

        # Build artifact metadata
        artifact_meta = {
            "columns": meta["columns"],
            "sheet_name": meta["sheet_name"],
            "sheet_names": meta["sheet_names"],
            "detected_lon_col": meta["lon_col"],
            "detected_lat_col": meta["lat_col"],
            "row_count": meta["row_count"],
            "valid_feature_count": meta["valid_feature_count"],
            "crs": crs if has_coords else None,
            "extent": meta["extent"],
        }

        if lazy:
            # Lazy mode: metadata only, LAZY_HANDLE backing
            artifact = Artifact(
                type=artifact_type,
                name=file_path.stem,
                backing=BackingStore(
                    kind=BackingStoreKind.LAZY_HANDLE,
                    uri=str(file_path),
                ),
                spatial=spatial,
                lineage=Lineage(operation="materialize", params=lineage_params),
                metadata=artifact_meta,
            )
            notes = (
                f"Excel metadata only — {meta['row_count']} rows from sheet '{meta['sheet_name']}'"
            )
            return MaterializeResult(
                artifact=artifact,
                strategy="lazy_handle",
                source_ref=source_ref,
                notes=notes,
            )

        # Eager mode
        if has_coords:
            # Write to GeoPackage
            output_path = workspace / f"{file_path.stem}.gpkg"
            try:
                _write_geopackage(file_path, output_path, meta, self.default_crs)
            except MaterializeError:
                raise
            except Exception as e:
                raise MaterializeError(source_ref, f"Failed to write GeoPackage: {e}") from e

            artifact = Artifact(
                type=ArtifactType.VECTOR,
                name=file_path.stem,
                backing=BackingStore(
                    kind=BackingStoreKind.LOCAL_FILE,
                    uri=str(output_path),
                    size_bytes=output_path.stat().st_size,
                    content_hash=content_hash(output_path),
                ),
                spatial=spatial,
                lineage=Lineage(operation="materialize", params=lineage_params),
                metadata=artifact_meta,
            )
            return MaterializeResult(
                artifact=artifact,
                strategy="wrapped_local",
                source_ref=source_ref,
                notes=f"Excel converted to GeoPackage ({meta['valid_feature_count']} features)",
            )
        else:
            # No coordinates: return as table with local file backing
            # For tables, we keep the original Excel file as backing
            artifact = Artifact(
                type=ArtifactType.TABLE,
                name=file_path.stem,
                backing=BackingStore(
                    kind=BackingStoreKind.LOCAL_FILE,
                    uri=str(file_path),
                    size_bytes=file_path.stat().st_size,
                    content_hash=content_hash(file_path),
                ),
                spatial=spatial,
                lineage=Lineage(operation="materialize", params=lineage_params),
                metadata=artifact_meta,
            )
            return MaterializeResult(
                artifact=artifact,
                strategy="wrapped_local",
                source_ref=source_ref,
                notes=f"Excel table (no coordinates detected) — {meta['row_count']} rows",
            )

    def discover(self, query: str | dict[str, Any] | None = None) -> list[CatalogEntry]:
        """List .xlsx/.xls files in a directory.

        Args:
            query: Directory path as string or dict with "path" key

        Returns:
            List of catalog entries for Excel files.
        """
        if isinstance(query, dict):
            dir_path = query.get("path")
        elif isinstance(query, str):
            dir_path = query
        else:
            raise MaterializeError("discover", "No path specified")

        if not dir_path:
            raise MaterializeError("discover", "No path specified")

        path = Path(dir_path)
        if not path.is_dir():
            raise MaterializeError("discover", f"Not a directory: {dir_path}")

        seen: set[str] = set()
        entries = []
        for pattern in ("*.xlsx", "*.XLSX", "*.xls", "*.XLS"):
            for file_path in path.glob(pattern):
                resolved = str(file_path.resolve())
                if resolved in seen:
                    continue
                seen.add(resolved)

                # Try to detect if this file has coordinates
                has_coords = False
                sheet_names = []
                try:
                    meta = _read_excel_metadata(file_path)
                    has_coords = meta["has_coordinates"]
                    sheet_names = meta["sheet_names"]
                except Exception:
                    pass

                entries.append(
                    CatalogEntry(
                        source_ref=str(file_path),
                        name=file_path.stem,
                        metadata={
                            "extension": file_path.suffix.lower(),
                            "has_coordinates": has_coords,
                            "sheet_names": sheet_names,
                        },
                    )
                )

        return entries

    def metadata(self, source_ref: SourceRef | str) -> dict[str, Any]:
        """Get Excel metadata without materializing data.

        Args:
            source_ref: Path to Excel file, optionally with "::sheet_name" or
                       "::sheet_name::lon_col,lat_col" suffix

        Returns:
            Metadata dict with columns, detected coordinates, row count, sheet names, etc.
        """
        file_path, sheet_name, explicit_lon_col, explicit_lat_col = _parse_source_ref(source_ref)

        try:
            meta = _read_excel_metadata(
                file_path,
                sheet_name=sheet_name,
                explicit_lon_col=explicit_lon_col,
                explicit_lat_col=explicit_lat_col,
            )
        except MaterializeError:
            raise
        except Exception as e:
            raise MaterializeError(source_ref, f"Failed to read Excel metadata: {e}") from e

        crs = meta["crs_hint"] if meta["crs_hint"] is not None else self.default_crs

        return {
            "columns": meta["columns"],
            "sheet_name": meta["sheet_name"],
            "sheet_names": meta["sheet_names"],
            "detected_lon_col": meta["lon_col"],
            "detected_lat_col": meta["lat_col"],
            "row_count": meta["row_count"],
            "valid_feature_count": meta["valid_feature_count"],
            "has_coordinates": meta["has_coordinates"],
            "crs": crs if meta["has_coordinates"] else None,
            "extent": meta["extent"],
        }
